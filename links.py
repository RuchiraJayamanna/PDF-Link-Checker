import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import urljoin, urlparse
from collections import deque
from openpyxl import load_workbook, Workbook
from multiprocessing import Pool


def get_pdf_links(url, filename):
    visited = set()
    queue = deque([url])
    pdf_issues = []

    while queue:
        current_url = queue.popleft()
        if current_url in visited:
            continue
        visited.add(current_url)

        try:
            response = requests.get(current_url)
            response.raise_for_status()
        except requests.exceptions.HTTPError as e:
            print(f"HTTPError: {e} for url: {current_url}")
            continue

        soup = BeautifulSoup(response.content, 'lxml')

        for link in soup.find_all('a', href=True):
            href = link['href']
            full_url = urljoin(current_url, href)

            if re.search(r'.*\.pdf$', href):
                print(f"PDF found: {full_url} on page: {current_url}")

                try:
                    pdf_response = requests.get(full_url)
                    pdf_response.raise_for_status()
                    update_excel(filename, (full_url, current_url), column='working')
                except requests.exceptions.RequestException as e:
                    print(f"Failed to load PDF: {full_url} from page: {current_url} with error: {e}")
                    pdf_issues.append((full_url, current_url))
                    update_excel(filename, (full_url, current_url), column='broken')
            else:
                if urlparse(full_url).netloc == urlparse(url).netloc and full_url not in visited:
                    queue.append(full_url)

    return pdf_issues


def initialize_excel(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = 'PDF Links'
    ws.append(['Working PDF Link', 'Found on Page', 'Broken PDF Link', 'Found on Page'])
    wb.save(filename)


def update_excel(filename, data, column):
    wb = load_workbook(filename)
    ws = wb['PDF Links']

    if column == 'working':
        ws.append([data[0], data[1], None, None])
    elif column == 'broken':
        ws.append([None, None, data[0], data[1]])

    wb.save(filename)


def process_url(args):
    url, filename = args
    get_pdf_links(url, filename)


def main():
    url = input("Plesae enter the URL to scrape for PDF links: ").strip()
    if not urlparse(url).scheme:
        url = 'http://' + url
    filename = 'pdf_links.xlsx'

    initialize_excel(filename)

    num_processes = 4
    urls = [(url, filename)] * num_processes

    with Pool(processes=num_processes) as pool:
        pool.map(process_url, urls)

    print(f"PDF links have been saved to '{filename}'.")


if __name__ == "__main__":
    main()